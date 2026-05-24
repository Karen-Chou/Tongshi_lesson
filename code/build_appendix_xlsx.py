from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from build_supporting_materials import OUT, industry_data, scenario_data, sources


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.font = Font(name="Microsoft YaHei", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


wb = Workbook()
ws = wb.active
ws.title = "行业应用指数"
ws.append(["行业", "AI渗透率", "投资强度", "数据可得性", "利润改善潜力", "社会影响权重", "综合应用指数", "典型场景"])
for row in industry_data:
    ws.append(row)
style_sheet(ws)
for col, width in enumerate([16, 12, 12, 12, 14, 14, 14, 34], start=1):
    ws.column_dimensions[get_column_letter(col)].width = width
ws.freeze_panes = "A2"
chart = BarChart()
chart.title = "行业AI综合应用指数"
chart.y_axis.title = "行业"
chart.x_axis.title = "指数"
data = Reference(ws, min_col=7, min_row=1, max_row=7)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 8
chart.width = 15
ws.add_chart(chart, "J2")

ws = wb.create_sheet("情景预测")
ws.append(["情景", "年份", "综合产出指数(2025=100)", "AI年生产率贡献百分点", "AI采用率假设", "人机协同任务占比"])
for row in scenario_data:
    ws.append(row)
style_sheet(ws)
for col, width in enumerate([14, 10, 22, 22, 16, 18], start=1):
    ws.column_dimensions[get_column_letter(col)].width = width
ws.freeze_panes = "A2"
line = LineChart()
line.title = "2030年前AI驱动产出指数情景"
line.y_axis.title = "产出指数"
line.x_axis.title = "年份"
for start in [2, 5, 8]:
    values = Reference(ws, min_col=3, min_row=start, max_row=start + 2)
    line.add_data(values, titles_from_data=False)
cats = Reference(ws, min_col=2, min_row=2, max_row=4)
line.set_categories(cats)
line.series[0].tx = SeriesLabel(v="保守情景")
line.series[1].tx = SeriesLabel(v="基准情景")
line.series[2].tx = SeriesLabel(v="积极情景")
line.height = 8
line.width = 15
ws.add_chart(line, "H2")

ws = wb.create_sheet("资料来源")
ws.append(["编号", "机构", "资料", "关键摘录", "链接"])
for row in sources:
    ws.append(row)
style_sheet(ws)
for col, width in enumerate([8, 18, 32, 72, 58], start=1):
    ws.column_dimensions[get_column_letter(col)].width = width
ws.freeze_panes = "A2"

ws = wb.create_sheet("模型说明")
rows = [
    ["模块", "说明"],
    ["应用成熟度指数", "0.25×渗透率 + 0.20×投资强度 + 0.20×数据可得性 + 0.20×利润改善潜力 + 0.15×社会影响权重"],
    ["随机森林", "以AI投入、数据质量、流程重构、人才结构等为特征，识别生产率和利润率变化的关键因素。"],
    ["MCMC", "围绕采用率、生产率贡献和人机协同任务占比设定先验分布，模拟未来路径不确定性。"],
    ["LP / CP-SAT", "在预算、算力、人才和风险约束下，最大化AI项目组合的经济与社会价值。"],
    ["注意", "本表为课程测算和展示用附录，不应表述为官方原始统计数据。"],
]
for row in rows:
    ws.append(row)
style_sheet(ws)
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 100

output = OUT / "数据与模型附录.xlsx"
wb.save(output)
print(output)
