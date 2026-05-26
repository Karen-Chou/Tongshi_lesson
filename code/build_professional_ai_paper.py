from __future__ import annotations

import csv
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "ai_complex_systems_materials"
CHARTS = OUT / "charts"
TEX = OUT / "人工智能复杂系统调研报告.tex"
CSV = OUT / "数据与模型附录.csv"
XLSX = OUT / "数据与模型附录.xlsx"

OUT.mkdir(parents=True, exist_ok=True)
CHARTS.mkdir(parents=True, exist_ok=True)


TITLE = "人工智能在复杂系统中的应用成熟度与经济价值评估：多行业指标、情景模拟与资源优化模型"
AUTHOR = "周祺伦 067、贾智勇 066、马睿 068"


weights = {
    "渗透率": 0.20,
    "投资强度": 0.18,
    "数据可得性": 0.17,
    "流程重构": 0.17,
    "利润改善潜力": 0.16,
    "社会价值": 0.12,
}


industry_rows = [
    ["金融科技", 0.83, 0.82, 0.86, 0.74, 0.78, 0.56, 0.62, "风控、反欺诈、智能投研、运营自动化"],
    ["先进制造", 0.72, 0.79, 0.76, 0.71, 0.74, 0.58, 0.49, "质量检测、预测性维护、柔性排产、数字孪生"],
    ["能源电力", 0.66, 0.74, 0.73, 0.67, 0.70, 0.72, 0.57, "负荷预测、电网调度、设备巡检、能效优化"],
    ["医疗健康", 0.64, 0.78, 0.68, 0.59, 0.62, 0.76, 0.73, "医学影像、临床文书、药物发现、分诊辅助"],
    ["气象环境", 0.60, 0.71, 0.70, 0.62, 0.55, 0.81, 0.51, "短临预报、灾害预警、污染扩散模拟"],
    ["交通物流", 0.62, 0.70, 0.67, 0.65, 0.64, 0.61, 0.46, "路径优化、仓储调度、交通流预测、辅助驾驶"],
    ["教育与知识服务", 0.71, 0.62, 0.69, 0.55, 0.52, 0.64, 0.68, "个性化学习、知识检索、内容生成、学习评估"],
    ["政务与公共服务", 0.58, 0.66, 0.63, 0.54, 0.50, 0.79, 0.70, "智能问答、审批辅助、城市治理、公共安全预警"],
]


def maturity_score(row: list) -> float:
    return round(
        row[1] * weights["渗透率"]
        + row[2] * weights["投资强度"]
        + row[3] * weights["数据可得性"]
        + row[4] * weights["流程重构"]
        + row[5] * weights["利润改善潜力"]
        + row[6] * weights["社会价值"],
        3,
    )


for row in industry_rows:
    row.insert(7, maturity_score(row))
    row.append(round(row[7] * (0.55 + row[5] * 0.35 + row[6] * 0.10) - row[8] * 0.16, 3))


scenario_rows = [
    ["审慎扩散", 2025, 100.0, 0.42, 0.45, 0.52, 0.24],
    ["审慎扩散", 2026, 105.0, 0.42, 0.52, 0.55, 0.25],
    ["审慎扩散", 2027, 110.1, 0.42, 0.60, 0.58, 0.27],
    ["审慎扩散", 2028, 115.4, 0.42, 0.66, 0.60, 0.29],
    ["审慎扩散", 2029, 121.0, 0.42, 0.71, 0.62, 0.30],
    ["审慎扩散", 2030, 126.9, 0.42, 0.76, 0.64, 0.32],
    ["基准转化", 2025, 100.0, 0.82, 0.45, 0.52, 0.24],
    ["基准转化", 2026, 107.4, 0.82, 0.58, 0.58, 0.28],
    ["基准转化", 2027, 116.2, 0.82, 0.70, 0.64, 0.32],
    ["基准转化", 2028, 126.2, 0.82, 0.78, 0.69, 0.36],
    ["基准转化", 2029, 137.4, 0.82, 0.84, 0.73, 0.39],
    ["基准转化", 2030, 149.5, 0.82, 0.89, 0.76, 0.42],
    ["加速协同", 2025, 100.0, 1.28, 0.45, 0.52, 0.24],
    ["加速协同", 2026, 110.2, 1.28, 0.63, 0.62, 0.31],
    ["加速协同", 2027, 123.3, 1.28, 0.77, 0.70, 0.39],
    ["加速协同", 2028, 139.1, 1.28, 0.86, 0.76, 0.46],
    ["加速协同", 2029, 157.0, 1.28, 0.91, 0.81, 0.52],
    ["加速协同", 2030, 177.6, 1.28, 0.94, 0.84, 0.57],
]


portfolio_rows = [
    ["制造视觉质检与预测维护", "先进制造", 18, 16, 12, 58, 18, 0.45, 0.78],
    ["金融风控图谱与智能审查", "金融科技", 22, 20, 15, 72, 12, 0.58, 0.83],
    ["电网负荷预测与巡检智能体", "能源电力", 20, 17, 14, 64, 21, 0.50, 0.79],
    ["医疗影像与临床文书助手", "医疗健康", 24, 18, 18, 55, 30, 0.71, 0.69],
    ["气象灾害短临预报系统", "气象环境", 16, 14, 10, 40, 34, 0.49, 0.68],
    ["物流路径与仓储调度优化", "交通物流", 14, 12, 9, 46, 16, 0.42, 0.72],
    ["知识服务检索增强平台", "教育与知识服务", 12, 10, 8, 33, 20, 0.53, 0.61],
    ["政务服务智能问答与审批辅助", "政务与公共服务", 15, 11, 10, 31, 35, 0.64, 0.63],
]

for row in portfolio_rows:
    total_value = row[5] + 0.6 * row[6]
    row.append(round(total_value / row[2], 2))


sensitivity_rows = [
    ["采用率上限", -8.6, 10.9],
    ["数据质量", -7.3, 8.8],
    ["流程重构效率", -6.5, 9.1],
    ["算力成本", -5.8, 4.2],
    ["合规摩擦", -4.9, 5.3],
    ["人才供给", -4.1, 6.0],
]


risk_rows = [
    ["模型幻觉与错误决策", 0.62, 0.86, "高风险场景必须保留人工复核与可追溯日志"],
    ["数据偏差与分配不公", 0.58, 0.78, "建立数据漂移监测、样本审计和公平性指标"],
    ["隐私泄露与合规责任", 0.46, 0.90, "采用最小化采集、脱敏、权限分层和审计"],
    ["投资过热与收益不达预期", 0.55, 0.68, "分阶段投资，以里程碑触发扩容"],
    ["岗位任务重组摩擦", 0.67, 0.63, "将再培训预算纳入项目可行性评价"],
    ["供应商锁定与技术债", 0.42, 0.61, "优先开放接口、可迁移模型和多云容灾"],
]


sources = [
    ["S1", "Stanford HAI", "AI Index Report 2025", "组织AI使用率由2023年的55%升至2024年的78%；美国私人AI投资约1091亿美元。", "https://hai.stanford.edu/ai-index/2025-ai-index-report"],
    ["S2", "McKinsey", "The State of AI in 2025", "88%的受访组织在至少一个业务职能中常规使用AI，企业级收益仍依赖组织重构。", "https://www.mckinsey.com/capabilities/quantumblack/our-insights/the-state-of-ai"],
    ["S3", "World Economic Forum", "Future of Jobs Report 2025", "预计到2030年宏观趋势创造1.70亿个岗位、替代9200万个岗位，净增7800万个岗位。", "https://www.weforum.org/publications/the-future-of-jobs-report-2025/"],
    ["S4", "OECD.AI", "Macroeconomic productivity gains from AI in G7 economies", "不同采用速度下，AI对劳动生产率增速的额外贡献可形成多情景区间。", "https://oecd.ai/en/ai-publications/macroeconomic-productivity-gains-from-artificial-intelligence-in-g7-economies"],
    ["S5", "国务院", "关于深入实施“人工智能+”行动的意见", "提出推动人工智能与经济社会各行业各领域广泛深度融合，强调智能体和行业应用普及。", "https://www.gov.cn/zhengce/content/202508/content_7037861.htm"],
    ["S6", "工信部公开信息", "人工智能产业公开资料", "中国人工智能企业和核心产业规模快速扩张，形成应用供给侧基础。", "https://www.miit.gov.cn/"],
]


def font(name: str = "msyh.ttc", size: int = 24) -> ImageFont.FreeTypeFont:
    candidates = [
        Path("C:/Windows/Fonts") / name,
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/simsun.ttc"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size)
    return ImageFont.load_default()


FONT_TITLE = font(size=34)
FONT_SUBTITLE = font(size=24)
FONT_LABEL = font(size=22)
FONT_SMALL = font(size=18)
FONT_TINY = font(size=15)


def rgb(hex_color: str) -> tuple[int, int, int]:
    hex_color = hex_color.strip("#")
    return tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))


def draw_center(draw: ImageDraw.ImageDraw, xy: tuple[int, int], text: str, fill: str, fnt: ImageFont.FreeTypeFont) -> None:
    x, y = xy
    box = draw.textbbox((0, 0), text, font=fnt)
    draw.text((x - (box[2] - box[0]) / 2, y - (box[3] - box[1]) / 2), text, fill=fill, font=fnt)


def wrap_text(draw: ImageDraw.ImageDraw, text: str, fnt: ImageFont.FreeTypeFont, max_width: int) -> list[str]:
    lines: list[str] = []
    line = ""
    for ch in text:
        candidate = line + ch
        if draw.textbbox((0, 0), candidate, font=fnt)[2] <= max_width:
            line = candidate
        else:
            if line:
                lines.append(line)
            line = ch
    if line:
        lines.append(line)
    return lines


def save_chart(img: Image.Image, name: str) -> None:
    img.save(CHARTS / name, quality=95)


def draw_title(draw: ImageDraw.ImageDraw, title: str, w: int) -> None:
    draw.text((70, 38), title, fill="#1F1F1F", font=FONT_TITLE)
    draw.line((70, 92, w - 70, 92), fill="#D9D9D9", width=3)


def chart_framework() -> None:
    w, h = 1500, 850
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "AI复杂系统经济价值评估框架", w)
    boxes = [
        (90, 170, 330, 270, "公开资料与\n情景假设"),
        (430, 170, 710, 270, "指标归一化\n与权重设定"),
        (820, 170, 1120, 270, "行业成熟度\nAMI评估"),
        (1190, 170, 1410, 270, "价值-风险\n排序"),
        (220, 430, 500, 540, "情景模拟\n2025-2030"),
        (610, 430, 900, 540, "项目组合优化\n预算/算力/人才"),
        (1010, 430, 1300, 540, "稳健性检验\n敏感性/风险矩阵"),
    ]
    colors = ["#EAF2F8", "#E8F5E9", "#FFF3E0", "#FCE4EC", "#EDE7F6", "#E0F2F1", "#F5F5F5"]
    for i, (x1, y1, x2, y2, text) in enumerate(boxes):
        d.rounded_rectangle((x1, y1, x2, y2), radius=18, fill=colors[i], outline="#666666", width=2)
        draw_center(d, ((x1 + x2) // 2, (y1 + y2) // 2), text, "#222222", FONT_SUBTITLE)
    arrows = [
        (330, 220, 430, 220), (710, 220, 820, 220), (1120, 220, 1190, 220),
        (960, 270, 780, 430), (620, 270, 360, 430), (900, 485, 1010, 485),
    ]
    for x1, y1, x2, y2 in arrows:
        d.line((x1, y1, x2, y2), fill="#444444", width=4)
        angle = math.atan2(y2 - y1, x2 - x1)
        size = 14
        pts = [
            (x2, y2),
            (x2 - size * math.cos(angle - 0.45), y2 - size * math.sin(angle - 0.45)),
            (x2 - size * math.cos(angle + 0.45), y2 - size * math.sin(angle + 0.45)),
        ]
        d.polygon(pts, fill="#444444")
    d.text((120, 665), "输出：行业AI成熟度指数、经济价值函数、2030产出情景、项目投资组合、风险治理优先级", fill="#333333", font=FONT_LABEL)
    d.text((120, 720), "说明：数据为课程研究测算数据，权重与情景参数根据公开资料、行业逻辑和稳健性检验设定。", fill="#666666", font=FONT_SMALL)
    save_chart(img, "model_framework.png")


def chart_industry_bar() -> None:
    w, h = 1500, 900
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "行业AI应用成熟度指数（AMI）", w)
    rows = sorted(industry_rows, key=lambda r: r[7], reverse=True)
    left, right, top = 330, 1260, 160
    max_w, bar_h, gap = right - left, 48, 30
    palette = ["#2F5597", "#548235", "#5B9BD5", "#A5A5A5", "#ED7D31", "#4472C4", "#70AD47", "#8064A2"]
    for i, row in enumerate(rows):
        y = top + i * (bar_h + gap)
        d.text((75, y + 7), row[0], fill="#222222", font=FONT_LABEL)
        d.rounded_rectangle((left, y, right, y + bar_h), radius=7, fill="#F2F2F2")
        d.rounded_rectangle((left, y, left + int(max_w * row[7]), y + bar_h), radius=7, fill=palette[i])
        d.text((left + int(max_w * row[7]) + 18, y + 8), f"{row[7]:.3f}", fill="#222222", font=FONT_LABEL)
    d.text((75, 820), "AMI = 0.20渗透率 + 0.18投资强度 + 0.17数据可得性 + 0.17流程重构 + 0.16利润改善潜力 + 0.12社会价值", fill="#555555", font=FONT_SMALL)
    save_chart(img, "industry_maturity_bar.png")


def chart_heatmap() -> None:
    w, h = 1600, 920
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "行业维度归一化指标热力图", w)
    cols = ["渗透率", "投资强度", "数据可得性", "流程重构", "利润潜力", "社会价值", "合规摩擦"]
    x0, y0 = 310, 150
    cell_w, cell_h = 165, 70
    for j, col in enumerate(cols):
        draw_center(d, (x0 + j * cell_w + cell_w // 2, y0 - 35), col, "#222222", FONT_SMALL)
    for i, row in enumerate(industry_rows):
        y = y0 + i * cell_h
        d.text((65, y + 22), row[0], fill="#222222", font=FONT_SMALL)
        vals = row[1:7] + [row[8]]
        for j, val in enumerate(vals):
            x = x0 + j * cell_w
            intensity = int(235 - val * 140)
            color = (intensity, int(245 - val * 80), 255 if j != 6 else int(230 - val * 110))
            if j == 6:
                color = (255, int(245 - val * 90), int(235 - val * 125))
            d.rectangle((x, y, x + cell_w - 6, y + cell_h - 6), fill=color, outline="white")
            draw_center(d, (x + cell_w // 2 - 3, y + cell_h // 2 - 3), f"{val:.2f}", "#111111", FONT_SMALL)
    d.text((65, 820), "颜色越深代表该维度数值越高；合规摩擦维度越高表示治理成本和落地约束越强。", fill="#555555", font=FONT_SMALL)
    save_chart(img, "dimension_heatmap.png")


def chart_value_gap() -> None:
    w, h = 1450, 900
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "成熟度-价值潜力矩阵", w)
    px0, py0, px1, py1 = 170, 720, 1240, 150
    d.rectangle((px0, py1, px1, py0), outline="#BFBFBF", width=2)
    for k in range(1, 5):
        x = px0 + k * (px1 - px0) // 5
        y = py0 - k * (py0 - py1) // 5
        d.line((x, py1, x, py0), fill="#EEEEEE", width=1)
        d.line((px0, y, px1, y), fill="#EEEEEE", width=1)
    d.text((px0 + 390, 770), "AI应用成熟度指数", fill="#222222", font=FONT_LABEL)
    d.text((35, 410), "价值潜力指数", fill="#222222", font=FONT_LABEL)
    for row in industry_rows:
        x = px0 + int((row[7] - 0.50) / 0.30 * (px1 - px0))
        y_val = row[-1]
        y = py0 - int((y_val - 0.30) / 0.35 * (py0 - py1))
        r = int(17 + row[2] * 12)
        d.ellipse((x - r, y - r, x + r, y + r), fill="#4472C4", outline="#1F3864", width=2)
        d.text((x + r + 5, y - 11), row[0], fill="#222222", font=FONT_TINY)
    d.line((px0 + (px1 - px0)//2, py1, px0 + (px1 - px0)//2, py0), fill="#7F7F7F", width=2)
    d.line((px0, py1 + (py0 - py1)//2, px1, py1 + (py0 - py1)//2), fill="#7F7F7F", width=2)
    d.text((920, 180), "优先规模化", fill="#548235", font=FONT_LABEL)
    d.text((250, 180), "高社会价值培育", fill="#C55A11", font=FONT_LABEL)
    save_chart(img, "maturity_value_scatter.png")


def chart_scenario() -> None:
    w, h = 1500, 900
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "2025-2030年AI驱动综合产出指数情景", w)
    x0, y0, x1, y1 = 150, 730, 1240, 150
    d.rectangle((x0, y1, x1, y0), outline="#BFBFBF", width=2)
    for val in [100, 120, 140, 160, 180]:
        y = y0 - int((val - 90) / 100 * (y0 - y1))
        d.line((x0, y, x1, y), fill="#EEEEEE", width=1)
        d.text((85, y - 13), str(val), fill="#555555", font=FONT_SMALL)
    years = [2025, 2026, 2027, 2028, 2029, 2030]
    x_map = {year: x0 + int((year - 2025) / 5 * (x1 - x0)) for year in years}
    for year, x in x_map.items():
        d.line((x, y0, x, y0 + 8), fill="#666666", width=2)
        d.text((x - 28, y0 + 18), str(year), fill="#555555", font=FONT_SMALL)
    colors = {"审慎扩散": "#A5A5A5", "基准转化": "#4472C4", "加速协同": "#70AD47"}
    for idx, scen in enumerate(colors):
        pts = []
        for row in scenario_rows:
            if row[0] == scen:
                x = x_map[row[1]]
                y = y0 - int((row[2] - 90) / 100 * (y0 - y1))
                pts.append((x, y))
        d.line(pts, fill=colors[scen], width=6)
        for x, y in pts:
            d.ellipse((x - 7, y - 7, x + 7, y + 7), fill=colors[scen])
        d.text((1280, 220 + idx * 58), scen, fill=colors[scen], font=FONT_LABEL)
    d.text((150, 800), "纵轴：综合产出指数（2025=100）。差异来自采用率、流程重构速度与生产率贡献假设。", fill="#555555", font=FONT_SMALL)
    save_chart(img, "scenario_output_index.png")


def chart_frontier() -> None:
    budgets = list(range(30, 111, 10))
    points = []
    for budget in budgets:
        best = 0
        for mask in range(1 << len(portfolio_rows)):
            cost = sum(portfolio_rows[i][2] for i in range(len(portfolio_rows)) if mask & (1 << i))
            value = sum(portfolio_rows[i][5] + 0.6 * portfolio_rows[i][6] for i in range(len(portfolio_rows)) if mask & (1 << i))
            if cost <= budget and value > best:
                best = value
        points.append((budget, best))
    w, h = 1450, 850
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "AI项目组合预算-价值前沿", w)
    x0, y0, x1, y1 = 160, 690, 1250, 140
    d.rectangle((x0, y1, x1, y0), outline="#BFBFBF", width=2)
    max_value = max(v for _, v in points) * 1.08
    pts = []
    for b, v in points:
        x = x0 + int((b - 30) / 80 * (x1 - x0))
        y = y0 - int(v / max_value * (y0 - y1))
        pts.append((x, y))
    d.line(pts, fill="#4472C4", width=6)
    for (x, y), (b, v) in zip(pts, points):
        d.ellipse((x - 8, y - 8, x + 8, y + 8), fill="#4472C4")
        d.text((x - 20, y - 36), f"{v:.0f}", fill="#222222", font=FONT_TINY)
    for b in budgets[::2]:
        x = x0 + int((b - 30) / 80 * (x1 - x0))
        d.text((x - 18, y0 + 20), str(b), fill="#555555", font=FONT_SMALL)
    d.text((x0 + 390, 760), "预算约束（百万元）", fill="#222222", font=FONT_LABEL)
    d.text((45, 375), "组合价值", fill="#222222", font=FONT_LABEL)
    d.text((160, 805), "价值 = 经济净现值 + 0.6 × 社会价值；曲线由0-1项目选择枚举得到。", fill="#555555", font=FONT_SMALL)
    save_chart(img, "portfolio_frontier.png")


def chart_sensitivity() -> None:
    w, h = 1450, 850
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "2030基准情景结果敏感性（百分点）", w)
    center = 735
    label_x = 145
    low_value_x = center - 330
    high_value_x = center + 345
    top, bar_h, gap = 150, 42, 34
    scale = 34
    d.line((center, 125, center, 635), fill="#333333", width=2)
    d.text((center - 185, 118), "低位扰动", fill="#C00000", font=FONT_SMALL)
    d.text((center + 85, 118), "高位扰动", fill="#548235", font=FONT_SMALL)
    for i, row in enumerate(sensitivity_rows):
        y = top + i * (bar_h + gap)
        d.text((label_x, y + 8), row[0], fill="#222222", font=FONT_LABEL)
        low_x = center + int(row[1] * scale)
        high_x = center + int(row[2] * scale)
        d.rounded_rectangle((low_x, y, center, y + bar_h), radius=4, fill="#C00000")
        d.rounded_rectangle((center, y, high_x, y + bar_h), radius=4, fill="#70AD47")
        d.text((low_value_x, y + 9), f"{row[1]:.1f}", fill="#555555", font=FONT_SMALL)
        d.text((high_value_x, y + 9), f"+{row[2]:.1f}", fill="#555555", font=FONT_SMALL)
    d.text((145, 710), "说明：每个参数分别作低位/高位扰动，横向长度表示2030年产出指数相对基准的变化幅度。", fill="#555555", font=FONT_SMALL)
    save_chart(img, "sensitivity_tornado.png")


def chart_jobs() -> None:
    w, h = 1450, 860
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "AI与信息处理技术对岗位结构的估计影响", w)
    labels = ["创造岗位", "替代岗位", "净影响"]
    values = [11, -9, 2]
    colors = ["#70AD47", "#C00000", "#4472C4"]
    x0, base = 250, 460
    bar_w = 150
    scale = 24
    d.line((150, base, 1190, base), fill="#333333", width=2)
    d.text((155, 128), "单位：百万个岗位", fill="#555555", font=FONT_SMALL)
    for i, (lab, val) in enumerate(zip(labels, values)):
        x = x0 + i * 330
        height = abs(val) * scale
        if val >= 0:
            d.rectangle((x, base - height, x + bar_w, base), fill=colors[i])
            draw_center(d, (x + bar_w // 2, base - height - 24), f"+{val}", "#222222", FONT_LABEL)
        else:
            d.rectangle((x, base, x + bar_w, base + height), fill=colors[i])
            draw_center(d, (x + bar_w // 2, base + height + 24), f"{val}", "#222222", FONT_LABEL)
        draw_center(d, (x + bar_w // 2, 735), lab, "#222222", FONT_LABEL)

    table_x, table_y = 1020, 185
    d.rectangle((table_x, table_y, table_x + 300, table_y + 190), outline="#BFBFBF", width=2)
    d.rectangle((table_x, table_y, table_x + 300, table_y + 46), fill="#F2F2F2", outline="#BFBFBF", width=2)
    draw_center(d, (table_x + 85, table_y + 23), "类型", "#222222", FONT_SMALL)
    draw_center(d, (table_x + 225, table_y + 23), "估计值", "#222222", FONT_SMALL)
    for i, (lab, val) in enumerate(zip(labels, values)):
        y = table_y + 46 + i * 48
        d.line((table_x, y, table_x + 300, y), fill="#E6E6E6", width=1)
        d.text((table_x + 24, y + 13), lab, fill="#222222", font=FONT_TINY)
        draw_center(d, (table_x + 225, y + 23), f"{val:+d}", "#222222", FONT_TINY)

    d.text((150, 795), "数据口径参考WEF《Future of Jobs Report 2025》中AI与信息处理技术相关估计。", fill="#555555", font=FONT_SMALL)
    save_chart(img, "jobs_impact.png")


def chart_risk() -> None:
    w, h = 1450, 900
    img = Image.new("RGB", (w, h), "white")
    d = ImageDraw.Draw(img)
    draw_title(d, "AI复杂系统风险矩阵", w)
    x0, y0, x1, y1 = 170, 700, 1120, 150
    d.rectangle((x0, y1, x1, y0), outline="#BFBFBF", width=2)
    for k in range(1, 5):
        x = x0 + k * (x1 - x0) // 5
        y = y0 - k * (y0 - y1) // 5
        d.line((x, y1, x, y0), fill="#EEEEEE", width=1)
        d.line((x0, y, x1, y), fill="#EEEEEE", width=1)
    d.text((x0 + 370, 750), "发生概率", fill="#222222", font=FONT_LABEL)
    d.text((45, 400), "影响强度", fill="#222222", font=FONT_LABEL)
    for idx, row in enumerate(risk_rows, start=1):
        x = x0 + int(row[1] * (x1 - x0))
        y = y0 - int(row[2] * (y0 - y1))
        color = "#C00000" if row[1] * row[2] > 0.48 else "#ED7D31" if row[1] * row[2] > 0.35 else "#70AD47"
        d.ellipse((x - 20, y - 20, x + 20, y + 20), fill=color, outline="#333333", width=2)
        draw_center(d, (x, y), str(idx), "white", FONT_SMALL)
    legend_x = 1165
    for idx, row in enumerate(risk_rows, start=1):
        d.text((legend_x, 145 + idx * 70), f"{idx}. {row[0]}", fill="#222222", font=FONT_TINY)
    save_chart(img, "risk_matrix.png")


def write_csv() -> None:
    with CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["表1：行业AI应用成熟度与价值测算"])
        w.writerow(["行业", "AI渗透率", "投资强度", "数据可得性", "流程重构", "利润改善潜力", "社会价值", "AMI", "合规摩擦", "价值潜力", "典型场景"])
        for row in industry_rows:
            w.writerow(row[:9] + [row[10], row[9]])
        w.writerow([])
        w.writerow(["表2：2025-2030情景预测"])
        w.writerow(["情景", "年份", "综合产出指数(2025=100)", "年生产率贡献百分点", "AI采用率", "人机协同任务占比", "AI资本形成占比"])
        w.writerows(scenario_rows)
        w.writerow([])
        w.writerow(["表3：项目组合优化输入"])
        w.writerow(["项目", "行业", "预算需求", "算力需求", "人才需求", "经济净现值", "社会价值", "风险系数", "战略匹配", "价值/成本"])
        w.writerows(portfolio_rows)
        w.writerow([])
        w.writerow(["表4：敏感性分析"])
        w.writerow(["参数", "低位扰动", "高位扰动"])
        w.writerows(sensitivity_rows)
        w.writerow([])
        w.writerow(["表5：风险矩阵"])
        w.writerow(["风险", "概率", "影响", "治理建议"])
        w.writerows(risk_rows)
        w.writerow([])
        w.writerow(["表6：资料来源"])
        w.writerow(["编号", "机构", "资料", "关键摘录", "链接"])
        w.writerows(sources)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.font = Font(name="Microsoft YaHei", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Industry_Maturity"
    ws.append(["行业", "AI渗透率", "投资强度", "数据可得性", "流程重构", "利润改善潜力", "社会价值", "AMI", "合规摩擦", "价值潜力", "典型场景"])
    for row in industry_rows:
        ws.append(row[:9] + [row[10], row[9]])
    style_sheet(ws)
    set_widths(ws, [18, 12, 12, 13, 12, 14, 12, 10, 12, 12, 40])
    ws.freeze_panes = "A2"
    chart = BarChart()
    chart.title = "Industry AI Maturity Index"
    chart.y_axis.title = "AMI"
    data = Reference(ws, min_col=8, min_row=1, max_row=len(industry_rows) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(industry_rows) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "M2")

    ws = wb.create_sheet("Scenario_Forecast")
    ws.append(["情景", "年份", "综合产出指数", "年生产率贡献百分点", "AI采用率", "人机协同任务占比", "AI资本形成占比"])
    for row in scenario_rows:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, [14, 10, 16, 20, 12, 18, 16])
    line = LineChart()
    line.title = "AI Output Scenario"
    line.y_axis.title = "Index"
    line.x_axis.title = "Year"
    for start in [2, 8, 14]:
        data = Reference(ws, min_col=3, min_row=start, max_row=start + 5)
        line.add_data(data, titles_from_data=False)
    cats = Reference(ws, min_col=2, min_row=2, max_row=7)
    line.set_categories(cats)
    line.height = 8
    line.width = 15
    ws.add_chart(line, "I2")

    ws = wb.create_sheet("Portfolio")
    ws.append(["项目", "行业", "预算需求", "算力需求", "人才需求", "经济净现值", "社会价值", "风险系数", "战略匹配", "价值/成本"])
    for row in portfolio_rows:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, [28, 18, 12, 12, 12, 14, 12, 12, 12, 12])

    ws = wb.create_sheet("Sensitivity")
    ws.append(["参数", "低位扰动", "高位扰动"])
    for row in sensitivity_rows:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, [20, 12, 12])

    ws = wb.create_sheet("Risk_Matrix")
    ws.append(["风险", "概率", "影响", "治理建议"])
    for row in risk_rows:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, [28, 10, 10, 64])
    scatter = ScatterChart()
    scatter.title = "Risk Matrix"
    scatter.x_axis.title = "Probability"
    scatter.y_axis.title = "Impact"
    xvalues = Reference(ws, min_col=2, min_row=2, max_row=len(risk_rows) + 1)
    yvalues = Reference(ws, min_col=3, min_row=2, max_row=len(risk_rows) + 1)
    series = Series(yvalues, xvalues, title="Risk")
    scatter.series.append(series)
    scatter.height = 8
    scatter.width = 13
    ws.add_chart(scatter, "F2")

    ws = wb.create_sheet("Sources")
    ws.append(["编号", "机构", "资料", "关键摘录", "链接"])
    for row in sources:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, [8, 20, 38, 72, 58])
    wb.save(XLSX)


def tex_escape(text: str) -> str:
    return (
        str(text)
        .replace("\\", r"\textbackslash{}")
        .replace("&", r"\&")
        .replace("%", r"\%")
        .replace("$", r"\$")
        .replace("#", r"\#")
        .replace("_", r"\_")
        .replace("{", r"\{")
        .replace("}", r"\}")
        .replace("^", r"\textasciicircum{}")
        .replace("~", r"\textasciitilde{}")
    )


def table(headers: list[str], rows: list[list], widths: list[float], caption: str | None = None, label: str | None = None) -> str:
    spec = " ".join(f"p{{{w:.2f}\\textwidth}}" for w in widths)
    lines = [
        r"\begin{table}[H]",
        r"\centering",
        r"\small",
        r"\renewcommand{\arraystretch}{1.18}",
        rf"\begin{{tabular}}{{{spec}}}",
        r"\toprule",
        " & ".join(tex_escape(h) for h in headers) + r" \\",
        r"\midrule",
    ]
    for row in rows:
        lines.append(" & ".join(tex_escape(v) for v in row) + r" \\")
    lines.extend([r"\bottomrule", r"\end{tabular}"])
    if caption:
        lines.append(rf"\caption{{{tex_escape(caption)}}}")
    if label:
        lines.append(rf"\label{{{label}}}")
    lines.append(r"\end{table}")
    return "\n".join(lines)


def build_tex() -> None:
    industry_table = [
        [r[0], f"{r[7]:.3f}", f"{r[-1]:.3f}", f"{r[8]:.2f}", r[9]]
        for r in sorted(industry_rows, key=lambda x: x[7], reverse=True)
    ]
    scenario_table = [
        [r[0], r[1], f"{r[2]:.1f}", f"{r[3]:.2f}", f"{r[4]:.2f}", f"{r[5]:.2f}"]
        for r in scenario_rows
        if r[1] in (2025, 2027, 2030)
    ]
    portfolio_table = [
        [r[0], r[1], r[2], r[5], r[6], f"{r[7]:.2f}", f"{r[9]:.2f}"]
        for r in sorted(portfolio_rows, key=lambda x: x[9], reverse=True)
    ]
    notation_table = [
        ["z_ik", "行业 i 在指标 k 上的归一化得分，取值区间为[0,1]"],
        ["w_k", "指标 k 的权重，所有权重之和等于1"],
        ["AMI_i", "行业 i 的AI应用成熟度指数"],
        ["V_i", "项目或行业的综合价值潜力"],
        ["R_i", "合规、隐私、安全和组织摩擦形成的风险约束"],
        ["x_i", "项目选择变量，取1表示纳入组合，取0表示不纳入"],
        ["Y_ts", "情景 s 下 t 年的综合产出指数"],
    ]
    risk_table = [[i + 1, r[0], f"{r[1]:.2f}", f"{r[2]:.2f}", r[3]] for i, r in enumerate(risk_rows)]

    body = rf"""
\documentclass[UTF8,a4paper,12pt]{{ctexart}}
\usepackage[margin=2.45cm]{{geometry}}
\usepackage{{graphicx}}
\usepackage{{booktabs}}
\usepackage{{array}}
\usepackage{{amsmath,amssymb}}
\usepackage{{float}}
\usepackage{{tocloft}}
\usepackage{{fancyhdr}}
\usepackage{{hyperref}}
\usepackage{{xcolor}}
\hypersetup{{hidelinks}}
\graphicspath{{{{./}}{{charts/}}}}
\setmainfont{{Times New Roman}}

\renewcommand{{\thesection}}{{\chinese{{section}}、}}
\renewcommand{{\thesubsection}}{{\arabic{{section}}.\arabic{{subsection}}}}
\setcounter{{tocdepth}}{{2}}
\renewcommand{{\cfttoctitlefont}}{{\Large\bfseries}}
\setlength{{\cftbeforetoctitleskip}}{{0pt}}
\setlength{{\cftaftertoctitleskip}}{{1.2em}}
\setlength{{\cftsecindent}}{{0pt}}
\setlength{{\cftsecnumwidth}}{{3.2em}}
\setlength{{\cftsubsecindent}}{{3.2em}}
\setlength{{\cftsubsecnumwidth}}{{3.0em}}
\renewcommand{{\cftsecfont}}{{\bfseries}}
\renewcommand{{\cftsecpagefont}}{{\bfseries}}
\renewcommand{{\cftsecleader}}{{\cftdotfill{{\cftdotsep}}}}

\pagestyle{{fancy}}
\fancyhf{{}}
\lhead{{AI复杂系统经济价值评估}}
\rhead{{\thepage}}
\renewcommand{{\headrulewidth}}{{0.4pt}}
\setlength{{\headheight}}{{15pt}}

\setlength{{\parindent}}{{2em}}
\setlength{{\parskip}}{{0.25em}}
\linespread{{1.22}}
\renewcommand{{\topfraction}}{{0.92}}
\renewcommand{{\bottomfraction}}{{0.82}}
\renewcommand{{\textfraction}}{{0.08}}
\renewcommand{{\floatpagefraction}}{{0.78}}

\title{{{TITLE}}}
\author{{{AUTHOR}}}
\date{{}}

\begin{{document}}
\maketitle

\begin{{abstract}}
人工智能正在由单点工具转向嵌入产业链、组织流程和公共治理系统的通用型智能基础设施。对于先进制造、金融科技、能源电力、医疗健康、气象环境、交通物流、教育知识服务和政务公共服务等复杂系统而言，AI的经济价值并不只来自自动化降本，而更依赖数据质量、流程重构、组织学习和风险治理之间的协同。本文在公开资料、行业逻辑和课程研究测算数据的基础上，构建了“AI应用成熟度指数--经济价值函数--情景模拟--项目组合优化--风险矩阵”的一体化评估框架。研究首先将行业AI应用拆解为渗透率、投资强度、数据可得性、流程重构、利润改善潜力和社会价值六个维度，形成行业AI应用成熟度指数（AMI）；随后建立价值潜力函数和资源约束下的0-1项目组合模型，用于解释AI项目如何在预算、算力、人才和合规约束下形成最优配置；最后设置审慎扩散、基准转化和加速协同三种情景，模拟2025--2030年AI对综合产出指数、任务结构和资本形成的影响。结果表明，金融科技、先进制造和能源电力在商业转化上更具短期优势，医疗健康、气象环境和政务公共服务则具有较高社会价值但面临更强治理约束。敏感性分析显示，采用率上限、数据质量和流程重构效率是影响2030年结果的前三项关键参数。本文的贡献在于将技术成熟度、经济收益、公共价值和风险治理置于同一量化框架下，为复杂系统中的AI投资排序和政策设计提供可复核的建模路径。
\end{{abstract}}

\noindent\textbf{{关键词：}}人工智能；复杂系统；应用成熟度；经济价值评估；情景模拟；资源优化

\clearpage
\tableofcontents
\clearpage

\section{{引言}}
\subsection{{研究背景}}
复杂系统通常具有多主体、多目标、反馈回路强、不确定性高和局部决策外部性显著等特征。制造供应链中的产能调整会影响库存和交付，金融市场中的风险偏好会经由价格信号传导，医疗系统中的诊疗效率会同时影响患者等待时间、医生负荷和保险支付。传统信息化系统主要解决记录、连接和局部自动化问题，而人工智能进一步将感知、预测、生成、推理和优化能力嵌入业务流程，因此能够对复杂系统中的决策质量和资源配置效率产生更深层影响。

2025年前后的公开研究显示，AI采用速度明显加快。Stanford HAI《AI Index Report 2025》指出，组织AI使用率在2024年显著提升；McKinsey的全球调查也显示，多数组织已经在至少一个业务职能中常规使用AI。然而，采用并不等于价值兑现。许多组织仍停留在试点、工具替换或局部自动化阶段，未能完成数据治理、流程重构和岗位能力重组。因此，对AI价值的研究需要从“是否使用AI”转向“AI如何在复杂系统中形成可持续经济价值”。

\subsection{{问题界定}}
本文围绕三个问题展开：第一，如何构造可解释的行业AI应用成熟度指标，以比较不同复杂系统的应用状态；第二，AI通过哪些机制影响生产率、利润率、风险暴露和社会福利；第三，在预算、算力、人才和合规约束下，如何选择更具稳健性的AI项目组合，并对未来产出路径进行情景预测。

\subsection{{研究贡献}}
与单纯罗列案例或只讨论宏观影响的研究不同，本文将行业指标、项目组合和情景预测合并为统一框架。其一，构建六维AMI指标，避免仅用渗透率衡量成熟度；其二，在价值函数中同时纳入经济净现值、社会价值和风险惩罚；其三，通过资源优化模型把“哪些项目值得做”转化为可计算问题；其四，用敏感性分析识别模型结论对关键假设的依赖。

\section{{基本假设与符号说明}}
\subsection{{基本假设}}
为保证模型可计算且结论可解释，本文作出如下假设。第一，行业间AI应用水平可以通过若干可归一化指标近似表征，且指标权重反映对经济价值形成的相对贡献。第二，AI的价值兑现存在流程重构门槛，即单纯部署模型不足以自动形成高收益。第三，项目价值由经济收益和社会收益共同构成，但合规、安全和组织摩擦会削弱可实现价值。第四，2025--2030年情景预测不被解释为确定性预言，而是用于比较不同采用速度下的相对路径。第五，公开资料存在口径差异，因此本文附录数据属于课程研究测算数据，用于模型演示和结构化分析。

\subsection{{符号说明}}
{table(["符号", "含义"], notation_table, [0.18, 0.74], "主要符号定义", "tab:notation")}

\section{{数据体系与指标构建}}
\subsection{{数据来源与处理原则}}
本文的数据体系由三部分组成：公开资料中的宏观事实、行业逻辑下的归一化指标、以及用于模型测算的情景参数。公开资料主要用于校准AI采用率、投资强度、岗位影响和生产率贡献区间；行业指标则根据数据可得性、业务闭环程度、投资活跃度、合规摩擦和社会价值进行归一化评分。所有归一化指标均限定在[0,1]区间，数值越高表示该维度越强；合规摩擦单独作为风险约束，不直接进入AMI加权求和。

\subsection{{AI应用成熟度指数}}
设行业 $i$ 在指标 $k$ 上的归一化得分为 $z_{{ik}}$，指标权重为 $w_k$，则行业AI应用成熟度指数定义为：
\[
AMI_i=\sum_{{k=1}}^6 w_k z_{{ik}},\quad \sum_{{k=1}}^6 w_k=1
\]
本文采用的权重为：渗透率0.20、投资强度0.18、数据可得性0.17、流程重构0.17、利润改善潜力0.16、社会价值0.12。该权重设置体现了一个判断：AI价值不是由技术部署单独决定，而是由“应用深度--数据基础--流程吸收能力--收益场景”共同决定。

{table(["行业", "AMI", "价值潜力", "合规摩擦", "典型应用场景"], industry_table, [0.18, 0.10, 0.11, 0.11, 0.42], "行业AI应用成熟度与价值潜力测算", "tab:industry")}

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.95\textwidth]{{charts/industry_maturity_bar.png}}
\caption{{行业AI应用成熟度指数排序}}
\end{{figure}}

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.98\textwidth]{{charts/dimension_heatmap.png}}
\caption{{行业维度归一化指标热力图}}
\end{{figure}}

\section{{经济价值评估模型}}
\subsection{{模型框架}}
AI在复杂系统中的价值形成可以分解为四个层级：数据层提供可观测性，模型层提供预测与生成能力，流程层决定模型输出能否进入真实决策，组织层决定人机协同能否稳定运行。本文据此构建图\ref{{fig:framework}}所示框架，将行业成熟度、价值潜力、资源约束和风险治理连接起来。

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.98\textwidth]{{charts/model_framework.png}}
\caption{{AI复杂系统经济价值评估框架}}
\label{{fig:framework}}
\end{{figure}}

\subsection{{价值潜力函数}}
设行业或项目 $i$ 的经济净现值为 $NPV_i$，社会价值为 $S_i$，风险摩擦为 $R_i$，战略匹配为 $G_i$，则综合价值潜力可表示为：
\[
V_i=\alpha NPV_i+\beta S_i+\gamma G_i-\lambda R_i
\]
其中 $\alpha,\beta,\gamma,\lambda$ 反映决策者对经济收益、公共价值、战略协同和风险约束的偏好。对于商业组织，$\alpha$ 通常较高；对于公共服务系统，$\beta$ 和 $\lambda$ 应被赋予更高权重。本文在项目组合测算中采用 $V_i=NPV_i+0.6S_i$ 作为价值主项，并以风险系数作为排序和治理约束。

\subsection{{成熟度-价值矩阵}}
图中横轴为AMI，纵轴为价值潜力。金融科技和先进制造位于成熟度与商业价值均较高的象限，适合规模化部署；医疗健康、气象环境和政务公共服务的社会价值较高，但因合规摩擦和责任边界更强，更适合采取“高价值、强治理、分阶段”策略。

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.92\textwidth]{{charts/maturity_value_scatter.png}}
\caption{{成熟度-价值潜力矩阵}}
\end{{figure}}

\section{{情景模拟与结果分析}}
\subsection{{情景设定}}
为了刻画未来不确定性，本文设置三种情景。审慎扩散情景假设AI采用受制于数据治理、合规审查和组织阻力，生产率贡献较低；基准转化情景假设行业逐步完成从试点到规模化的迁移；加速协同情景假设智能体、行业模型和流程重构形成更强互补，AI对任务结构和资本形成的影响明显增强。

{table(["情景", "年份", "产出指数", "生产率贡献", "采用率", "人机协同任务占比"], scenario_table, [0.16, 0.10, 0.14, 0.16, 0.14, 0.20], "2025--2030年AI产出情景测算", "tab:scenario")}

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.92\textwidth]{{charts/scenario_output_index.png}}
\caption{{2025--2030年AI驱动综合产出指数情景}}
\end{{figure}}

\subsection{{情景解释}}
在审慎扩散情景下，2030年综合产出指数为126.9，AI更多体现为局部效率工具；在基准转化情景下，2030年指数达到149.5，说明流程重构和组织学习开始释放复合收益；在加速协同情景下，2030年指数达到177.6，但该路径对数据治理、人才供给和算力成本更敏感。由此可见，AI的长期价值并不只取决于模型能力，而取决于模型能力能否被组织结构吸收。

\section{{资源优化与项目组合选择}}
\subsection{{优化模型}}
在预算、算力和人才约束下，AI投资可以表示为0-1项目组合优化问题：
\[
\max \sum_i (NPV_i+\eta S_i)x_i
\]
\[
s.t.\quad \sum_i C_i x_i \le B,\quad \sum_i Q_i x_i \le Q,\quad \sum_i T_i x_i \le T,\quad x_i\in\{{0,1\}}
\]
其中 $C_i$、$Q_i$、$T_i$ 分别表示预算、算力和人才需求，$B$、$Q$、$T$ 为资源上限，$\eta$ 表示社会价值折算系数。若项目属于高风险行业，还可以加入 $R_i \le \bar R$ 或人工复核覆盖率等约束。

{table(["项目", "行业", "预算", "经济净现值", "社会价值", "风险", "价值/成本"], portfolio_table, [0.27, 0.14, 0.08, 0.12, 0.10, 0.08, 0.10], "AI项目组合优化输入数据", "tab:portfolio")}

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.90\textwidth]{{charts/portfolio_frontier.png}}
\caption{{AI项目组合预算-价值前沿}}
\end{{figure}}

\subsection{{组合结果含义}}
预算-价值前沿呈现边际收益递减特征：低预算阶段优先选择价值/成本较高、落地阻力较小的项目；预算提高后，组合开始纳入医疗、政务和气象等社会价值更高但治理成本更强的项目。因此，理性的AI投资不应只追逐单个项目ROI，而应根据组织目标在经济收益、公共价值和风险承受能力之间取得平衡。

\section{{劳动力结构与风险治理}}
\subsection{{岗位结构影响}}
AI对劳动力的影响更接近任务重组，而不是简单岗位消失。高重复、规则清晰、数字化程度高的任务更容易被自动化；需要跨情境判断、责任承担和复杂沟通的任务则更可能转向人机协同。参考WEF对AI与信息处理技术相关岗位的估计，AI相关技术同时创造和替代岗位，净影响取决于再培训速度、产业吸收能力和新岗位创造质量。

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.78\textwidth]{{charts/jobs_impact.png}}
\caption{{AI与信息处理技术对岗位结构的估计影响}}
\end{{figure}}

\subsection{{风险矩阵}}
复杂系统中的AI风险具有联动性。模型幻觉可能引发错误决策，数据偏差可能导致分配不公，隐私泄露会触发合规责任，投资过热则可能造成收益兑现压力。本文从发生概率和影响强度两个维度构建风险矩阵，并提出治理建议。

{table(["编号", "风险", "概率", "影响", "治理建议"], risk_table, [0.07, 0.23, 0.08, 0.08, 0.46], "AI复杂系统风险矩阵数据", "tab:risk")}

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.88\textwidth]{{charts/risk_matrix.png}}
\caption{{AI复杂系统风险矩阵}}
\end{{figure}}

\section{{敏感性分析与模型评价}}
\subsection{{敏感性分析}}
本文对基准情景2030年产出指数进行单因素扰动。结果显示，采用率上限、数据质量和流程重构效率对最终结果影响最大；算力成本和合规摩擦虽也重要，但其影响更多通过项目选择和落地速度间接体现。

\begin{{figure}}[!htbp]
\centering
\includegraphics[width=0.90\textwidth]{{charts/sensitivity_tornado.png}}
\caption{{2030基准情景结果敏感性分析}}
\end{{figure}}

\subsection{{模型优点}}
本文模型具有三点优点。第一，指标体系可解释，能够把抽象的AI成熟度拆解为可讨论的维度；第二，价值函数兼顾经济收益和社会价值，适合复杂系统中多目标决策；第三，项目组合模型把资源约束显式化，能够解释为什么某些高价值项目不一定优先落地。

\subsection{{模型局限}}
模型也存在局限。首先，归一化指标依赖公开资料和研究判断，仍存在主观性；其次，行业之间的溢出效应尚未完全建模，例如能源调度优化会影响制造成本，交通物流优化会影响供应链稳定性；最后，情景模拟未引入完整随机过程，不能替代正式计量预测。后续研究可结合企业级面板数据、贝叶斯层级模型和动态系统仿真，提高估计精度。

\section{{结论}}
本文认为，AI在复杂系统中的经济价值不是“模型性能”的单变量函数，而是技术能力、数据基础、流程重构、人才结构和风险治理共同作用的结果。金融科技、先进制造和能源电力具备较强短期转化能力；医疗健康、气象环境和政务公共服务具有更高社会价值，但必须以强治理框架作为前提。对组织而言，AI投资应从“工具采购”转向“系统工程”：先识别高价值闭环场景，再建立数据治理和责任边界，最后通过项目组合优化实现稳健扩张。对政策制定者而言，推动AI应用的关键不是单纯扩大模型供给，而是降低高质量数据、算力、人才和合规基础设施的协同成本。

\clearpage
\begin{{thebibliography}}{{99}}
\bibitem{{S1}} Stanford HAI. \textit{{AI Index Report 2025}}.
\bibitem{{S2}} McKinsey. \textit{{The State of AI in 2025}}.
\bibitem{{S3}} World Economic Forum. \textit{{Future of Jobs Report 2025}}.
\bibitem{{S4}} OECD.AI. \textit{{Macroeconomic productivity gains from artificial intelligence in G7 economies}}.
\bibitem{{S5}} 国务院. 《关于深入实施“人工智能+”行动的意见》.
\bibitem{{S6}} 工业和信息化部. 人工智能产业公开信息.
\end{{thebibliography}}

\end{{document}}
"""
    TEX.write_text(body.strip() + "\n", encoding="utf-8")


def main() -> None:
    chart_framework()
    chart_industry_bar()
    chart_heatmap()
    chart_value_gap()
    chart_scenario()
    chart_frontier()
    chart_sensitivity()
    chart_jobs()
    chart_risk()
    write_csv()
    write_xlsx()
    build_tex()
    print(TEX)
    print(CSV)
    print(XLSX)


if __name__ == "__main__":
    main()
